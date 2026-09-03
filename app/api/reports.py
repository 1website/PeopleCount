import io
import re
import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Family, Member, Village, Commune, District, Province
from app.auth import require_user

router = APIRouter(prefix="/api/reports", tags=["Reporting and Analytics"])

KHMER_TO_LATIN_DIGITS = {
    '០': '0', '១': '1', '២': '2', '៣': '3', '៤': '4',
    '៥': '5', '៦': '6', '៧': '7', '៨': '8', '៩': '9'
}

def khmer_to_latin_digits(text: Optional[str]) -> Optional[str]:
    if not text:
        return text
    res = str(text)
    for kh, lat in KHMER_TO_LATIN_DIGITS.items():
        res = res.replace(kh, lat)
    return res


@router.get("/dashboard-stats")
def get_dashboard_stats(
    village_id: Optional[int] = None,
    commune_id: Optional[int] = None,
    district_id: Optional[int] = None,
    province_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Computes overall statistics:
    - Total population (Male/Female)
    - Birth certificates (Have/None)
    - School age children (6-11, 12-14, 15-17)
    - School attendance & dropout counts/rates
    - Poverty status (IDPoor 1, IDPoor 2, General)
    """
    fam_query = db.query(Family)
    
    if village_id:
        fam_query = fam_query.filter(Family.village_id == village_id)
    elif commune_id:
        v_ids = [v.id for v in db.query(Village.id).filter(Village.commune_id == commune_id).all()]
        fam_query = fam_query.filter(Family.village_id.in_(v_ids))
    elif district_id:
        c_ids = [c.id for c in db.query(Commune.id).filter(Commune.district_id == district_id).all()]
        v_ids = [v.id for v in db.query(Village.id).filter(Village.commune_id.in_(c_ids)).all()]
        fam_query = fam_query.filter(Family.village_id.in_(v_ids))
    elif province_id:
        d_ids = [d.id for d in db.query(District.id).filter(District.province_id == province_id).all()]
        c_ids = [c.id for c in db.query(Commune.id).filter(Commune.district_id.in_(d_ids)).all()]
        v_ids = [v.id for v in db.query(Village.id).filter(Village.commune_id.in_(c_ids)).all()]
        fam_query = fam_query.filter(Family.village_id.in_(v_ids))

    families = fam_query.all()
    fam_ids = [f.id for f in families]

    total_families = len(families)
    poor_1_count = sum(1 for f in families if f.poor_category == "IDPOOR_1")
    poor_2_count = sum(1 for f in families if f.poor_category == "IDPOOR_2")
    general_count = sum(1 for f in families if f.poor_category == "GENERAL")
    pending_review = sum(1 for f in families if f.status == "PENDING_REVIEW")
    approved_count = sum(1 for f in families if f.status == "APPROVED")

    # Member-level stats
    if not fam_ids:
        members = []
    else:
        members = db.query(Member).filter(Member.family_id.in_(fam_ids)).all()

    total_population = len(members)
    male_count = sum(1 for m in members if m.gender == "MALE")
    female_count = sum(1 for m in members if m.gender == "FEMALE")

    birth_cert_yes = sum(1 for m in members if m.birth_cert and str(m.birth_cert).strip() not in ["0", "", "None", "False"])
    birth_cert_no = total_population - birth_cert_yes

    # Children and School-age Breakdown (including Infants age 0)
    infants_age_0 = [m for m in members if m.age == 0]
    toddlers_age_1_2 = [m for m in members if 1 <= m.age <= 2]
    kindergarten_age = [m for m in members if 3 <= m.age <= 5]
    primary_age = [m for m in members if 6 <= m.age <= 11]
    lower_sec_age = [m for m in members if 12 <= m.age <= 14]
    upper_sec_age = [m for m in members if 15 <= m.age <= 17]
    total_school_age = len(primary_age) + len(lower_sec_age) + len(upper_sec_age)
    total_children = len([m for m in members if m.age < 18])

    # Dropout statistics (across school-age children)
    all_school_youth = primary_age + lower_sec_age + upper_sec_age
    school_age_dropouts = sum(1 for m in all_school_youth if m.dropout_status == "DROPOUT")
    school_age_suspended = sum(1 for m in all_school_youth if m.dropout_status == "SUSPENDED")
    dropout_rate = (
        round((school_age_dropouts / len(all_school_youth)) * 100, 1)
        if all_school_youth else 0.0
    )

    # Education statuses across all population
    edu_none = sum(1 for m in members if m.education_status == "NONE")
    edu_primary = sum(1 for m in members if m.education_status == "PRIMARY")
    edu_secondary = sum(1 for m in members if m.education_status == "SECONDARY")
    edu_higher = sum(1 for m in members if m.education_status == "HIGHER")

    # Dropouts breakdown by grade and groups (0-6, 7-9, 10-12)
    dropout_breakdown = {}
    dropout_groups = {
        "grades_0_6": 0,
        "grades_7_9": 0,
        "grades_10_12": 0,
        "other": 0
    }
    for m in members:
        if m.dropout_status == "DROPOUT":
            if m.dropout_grade:
                norm_grade = khmer_to_latin_digits(m.dropout_grade).strip()
                dropout_breakdown[norm_grade] = dropout_breakdown.get(norm_grade, 0) + 1

                # Extract numeric grade
                match = re.search(r'\d+', norm_grade)
                if match:
                    grade_num = int(match.group())
                    if 0 <= grade_num <= 6:
                        dropout_groups["grades_0_6"] += 1
                    elif 7 <= grade_num <= 9:
                        dropout_groups["grades_7_9"] += 1
                    elif 10 <= grade_num <= 12:
                        dropout_groups["grades_10_12"] += 1
                    else:
                        dropout_groups["other"] += 1
                else:
                    dropout_groups["other"] += 1
            else:
                dropout_groups["other"] += 1

    return {
        "families": {
            "total": total_families,
            "poor_1": poor_1_count,
            "poor_2": poor_2_count,
            "general": general_count,
            "pending_review": pending_review,
            "approved": approved_count
        },
        "demographics": {
            "total_population": total_population,
            "male": male_count,
            "female": female_count,
            "female_percentage": round((female_count / total_population * 100), 1) if total_population else 0.0
        },
        "birth_certificate": {
            "have": birth_cert_yes,
            "none": birth_cert_no,
            "percentage_have": round((birth_cert_yes / total_population * 100), 1) if total_population else 0.0
        },
        "education": {
            "infants_0": len(infants_age_0),
            "toddlers_1_2": len(toddlers_age_1_2),
            "kindergarten_3_5": len(kindergarten_age),
            "primary_6_11": len(primary_age),
            "lower_sec_12_14": len(lower_sec_age),
            "upper_sec_15_17": len(upper_sec_age),
            "total_school_age": total_school_age,
            "total_children": total_children,
            "dropouts_count": school_age_dropouts,
            "suspended_count": school_age_suspended,
            "dropout_rate_percent": dropout_rate,
            "status_distribution": {
                "none": edu_none,
                "primary": edu_primary,
                "secondary": edu_secondary,
                "higher": edu_higher
            },
            "dropout_breakdown": dropout_breakdown,
            "dropout_groups": dropout_groups
        }
    }


@router.get("/export/excel")
def export_excel_report(
    village_id: Optional[int] = None,
    commune_id: Optional[int] = None,
    district_id: Optional[int] = None,
    province_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Exports population and family statistics to an official Excel format (.xlsx)
    with Cambodian administrative formatting, colors, borders, and clear headers.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ស្ថិតិប្រជាជន និងគ្រួសារ"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_national_header = Font(name="Khmer OS Muol Light", size=14, bold=True, color="001F3F")
    font_sub_header = Font(name="Khmer OS Battambang", size=11, bold=True, color="1E3A8A")
    font_table_header = Font(name="Khmer OS Battambang", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Khmer OS Battambang", size=10)
    font_bold_data = Font(name="Khmer OS Battambang", size=10, bold=True)

    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
    fill_sub_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_even_row = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Header section (National Motto)
    ws.merge_cells("A1:N1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = font_national_header
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:N2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = font_sub_header
    ws["A2"].alignment = align_center

    # Title
    ws.merge_cells("A4:N4")
    loc_title = "តារាងស្រង់ស្ថិតិប្រជាជន និងស្ថានភាពគ្រួសារលម្អិតតាមមូលដ្ឋាន"
    if village_id:
        v = db.query(Village).filter(Village.id == village_id).first()
        if v:
            loc_title += f" (ភូមិ {v.name_kh} - កូដ {v.code})"
    elif commune_id:
        c = db.query(Commune).filter(Commune.id == commune_id).first()
        if c:
            loc_title += f" (ឃុំ/សង្កាត់ {c.name_kh} - កូដ {c.code})"
    elif district_id:
        d = db.query(District).filter(District.id == district_id).first()
        if d:
            loc_title += f" (ស្រុក/ខណ្ឌ {d.name_kh} - កូដ {d.code})"
    elif province_id:
        p = db.query(Province).filter(Province.id == province_id).first()
        if p:
            loc_title += f" ({p.name_kh} - កូដ {p.code})"
    ws["A4"] = loc_title
    ws["A4"].font = Font(name="Khmer OS Muol Light", size=12, bold=True, color="0F172A")
    ws["A4"].alignment = align_center

    ws["A5"] = f"កាលបរិច្ឆេទចេញរបាយការណ៍៖ {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A5"].font = Font(name="Khmer OS Battambang", size=9, italic=True)

    # Table Column Headers
    headers = [
        "ល.រ",
        "លេខកូដគ្រួសារ",
        "ប្រភេទគ្រួសារ",
        "គោត្តនាម និងនាម",
        "ភេទ",
        "សញ្ជាតិ",
        "ថ្ងៃខែឆ្នាំកំណើត",
        "អាយុ",
        "ឋានៈក្នុងគ្រួសារ",
        "កម្រិតវប្បធម៌",
        "ស្ថានភាពសិក្សា",
        "សំបុត្រកំណើត",
        "មុខរបរ",
        "ទីតាំងភូមិ/ឃុំ"
    ]

    header_row = 7
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h_text)
        cell.font = font_table_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Fetch data
    fam_query = db.query(Family)
    if village_id:
        fam_query = fam_query.filter(Family.village_id == village_id)
    elif commune_id:
        v_ids = [v.id for v in db.query(Village.id).filter(Village.commune_id == commune_id).all()]
        fam_query = fam_query.filter(Family.village_id.in_(v_ids))
    elif district_id:
        c_ids = [c.id for c in db.query(Commune.id).filter(Commune.district_id == district_id).all()]
        v_ids = [v.id for v in db.query(Village.id).filter(Village.commune_id.in_(c_ids)).all()]
        fam_query = fam_query.filter(Family.village_id.in_(v_ids))
    elif province_id:
        d_ids = [d.id for d in db.query(District.id).filter(District.province_id == province_id).all()]
        c_ids = [c.id for c in db.query(Commune.id).filter(Commune.district_id.in_(d_ids)).all()]
        v_ids = [v.id for v in db.query(Village.id).filter(Village.commune_id.in_(c_ids)).all()]
        fam_query = fam_query.filter(Family.village_id.in_(v_ids))

    families = fam_query.order_by(Family.id.asc()).all()

    row_idx = 8
    seq = 1

    poor_map = {"IDPOOR_1": "ក្រ១", "IDPOOR_2": "ក្រ២", "GENERAL": "ទូទៅ"}
    relation_map = {
        "HEAD": "មេគ្រួសារ", "SPOUSE": "ប្រពន្ធ/ប្តី", "CHILD": "កូន",
        "PARENT": "ឪពុក/ម្តាយ", "RELATIVE": "សាច់ញាតិ", "OTHER": "ផ្សេងៗ"
    }
    edu_map = {"NONE": "មិនបានរៀន", "PRIMARY": "បឋមសិក្សា", "SECONDARY": "មធ្យមសិក្សា", "HIGHER": "ឧត្តមសិក្សា"}
    dropout_map = {"ACTIVE": "កំពុងរៀន", "DROPOUT": "បោះបង់", "SUSPENDED": "បង្អង់", "COMPLETED": "បានបញ្ចប់"}

    for fam in families:
        vill_name = f"{fam.village.name_kh} ({fam.village.commune.name_kh})" if fam.village else "N/A"
        fam_poor = poor_map.get(fam.poor_category, fam.poor_category)

        for member in fam.members:
            gender_kh = "ប្រុស" if member.gender == "MALE" else "ស្រី"
            birth_cert_val = str(member.birth_cert).strip() if member.birth_cert is not None else "0"
            birth_cert_kh = birth_cert_val if birth_cert_val not in ["0", "False", "None", ""] else "0 (គ្មាន)"
            edu_kh = edu_map.get(member.education_status, member.education_status)
            if member.dropout_status == "DROPOUT" and member.dropout_grade:
                study_status_kh = f"បោះបង់ ({khmer_to_latin_digits(member.dropout_grade)})"
            elif member.dropout_status == "COMPLETED" and member.dropout_grade:
                study_status_kh = f"បានបញ្ចប់ ({khmer_to_latin_digits(member.dropout_grade)})"
            elif member.dropout_status == "ACTIVE" and member.dropout_grade:
                study_status_kh = f"កំពុងរៀន ({khmer_to_latin_digits(member.dropout_grade)})"
            else:
                study_status_kh = dropout_map.get(member.dropout_status, member.dropout_status)

            values = [
                seq,
                fam.family_code,
                fam_poor,
                member.full_name,
                gender_kh,
                member.nationality,
                member.dob.strftime("%d/%m/%Y"),
                member.age,
                relation_map.get(member.relation, member.relation),
                edu_kh,
                study_status_kh,
                birth_cert_kh,
                member.occupation or "-",
                vill_name
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                if col_idx in [1, 5, 7, 8, 9, 12]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

                if row_idx % 2 == 0:
                    cell.fill = fill_even_row

            row_idx += 1
            seq += 1

    # Auto adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len and cell.row > 5:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Signatures block at bottom
    row_idx += 2
    ws.cell(row=row_idx, column=10, value="ធ្វើនៅ..................., ថ្ងៃទី..... ខែ..... ឆ្នាំ ២០២...")
    ws.cell(row=row_idx, column=10).font = font_bold_data
    row_idx += 1
    ws.cell(row=row_idx, column=2, value="អ្នកស្រង់ស្ថិតិ").font = font_bold_data
    ws.cell(row=row_idx, column=6, value="មេភូមិ បានឃើញ និងបញ្ជាក់").font = font_bold_data
    ws.cell(row=row_idx, column=11, value="ប្រធានក្រុមប្រឹក្សាឃុំ/សង្កាត់").font = font_bold_data

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Cambodia_Census_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
